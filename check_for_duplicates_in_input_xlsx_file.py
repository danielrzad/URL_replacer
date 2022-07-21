from pprint import pprint
from openpyxl import load_workbook
from openpyxl.styles import NamedStyle
from openpyxl.worksheet.cell_range import CellRange
from ordered_set import OrderedSet
from urllib.parse import urlparse
from dataclasses import make_dataclass


import yaml
import os
import webbrowser
import time


#### TO DO 
#   Dopisać żeby nie traktował placeholderów jako duplikaty!!!!!


def get_shortened_url(url):
    ### This function is cutting http/s in front of url to be sure there won't be http/https missmatches
    ### and .com fragment at the end of the netloc in case a website changes their .extension  
    urlparse("scheme://netloc/path;parameters?query#fragment")
    parsed_url = urlparse(url)
    netlocc = parsed_url.netloc[:parsed_url.netloc.rfind('.')].replace('www.', '')
    pathh = parsed_url.path
    #parameterss = parsed_url.parameters
    queryy = parsed_url.query
    fragmentt = parsed_url.fragment
    return (netlocc + pathh + queryy + fragmentt)


def find_duplicated_urls(input_ws, place_holder):
    ### Add all duplicates to a dict with scheme like:
    ###     {url: [1st_DataClass_with_url_data, 2nd_DataClass_with_url_data, etc]}
    ###
    ContainerAttributes = make_dataclass(
        'ContainerAttributes', ['website_url', 'date_found', 'container_url', 'title', 'cell_row']
    )
    urls_seen = {}
    for cell in input_ws['c4':f'z{input_ws.max_row}']:
        container_attributes = ContainerAttributes(
                                    website_url=cell[0].value,
                                    date_found=cell[6].value,
                                    container_url=cell[9].value,
                                    title=cell[23].value,
                                    cell_row=cell[9].row,
                                )
        if container_attributes.container_url == place_holder:
            continue
        url_stripped = get_shortened_url(container_attributes.container_url)
        if url_stripped not in urls_seen:
            urls_seen[url_stripped] = []
            urls_seen[url_stripped].append(container_attributes)
            continue
        urls_seen[url_stripped].append(container_attributes)
    duplicates = {key: value for key, value in urls_seen.items() if len(value) > 1}
    return duplicates


def create_list_for_fast_check(duplicates):
    duplicates_to_fast_check = []
    for key, value in duplicates.items():
        value_ordered = []
        for idx, val in enumerate(value):
            value_ordered.append([val, idx, val.date_found])
        value_ordered.sort(key = lambda x: x[2])
        value_ordered = [x[0] for x in value_ordered]
        duplicates_to_fast_check.append(value_ordered)
    return duplicates_to_fast_check


def fast_urls_check(duplicates_ordered_by_date_found, pause_between_opening_urls):
    value = duplicates_ordered_by_date_found
    ### takes ContainerAttributes dataclass to a duplicated container url
    #   opens title of the first duplicate
    #   then opens correlated container urls in chronological order
    #   (from older to newer)
    chrome_path = 'C:/Program Files/Google/Chrome/Application/chrome.exe %s'
    rows_to_delete = []
    wait_for_user_check = 0
    for duplicate_title in value:
        title_formated = duplicate_title[0].title.replace(' ', '-').replace(':', '-')
        webbrowser.get(chrome_path).open(f'www.zxdsxzxs.pl/{title_formated}')
        for url in duplicate_title:
            webbrowser.get(chrome_path).open(url.container_url)
            time.sleep(pause_between_opening_urls)
            rows_to_delete.append(url.cell_row)
            wait_for_user_check += 1
        if wait_for_user_check >= 8:
            wait_for_user_check = 0
            input('Check opened URLs and press "any_key" to continue:\n')
        ### line below deletes last element of the rows_to_delete
        #   because it's the newest and thus valid row
        #   coresponding to an container url which we want to keep
        del rows_to_delete[-1]
    ask_for_permision_to_delete = input(
        'Are all of the newest URLs correct?\n'
        f'Those rows will be deleted:\n{rows_to_delete}\n'
        'Enter "y" for deleting and "n" for pausing execution and moving to manual selection.\n'
    )
    print()
    if ask_for_permision_to_delete == 'y':
        return rows_to_delete
    print('Some new duplicated URLs are not correct.\nProgram will proceed to a "manual_urls_check".\n')
    return []


def manual_urls_check(duplicates, pause_between_opening_urls):
    chrome_path = 'C:/Program Files/Google/Chrome/Application/chrome.exe %s'
    rows_to_delete = []
    for key, value in duplicates.items():
        ### Get single title with container urls sorted by date
        duplicates_ordered_by_date = []
        for idx, val in enumerate(value):
            duplicates_ordered_by_date.append([val, idx, val.date_found])
        duplicates_ordered_by_date.sort(key = lambda x: x[2])
        duplicates_ordered_by_date = [x[0] for x in duplicates_ordered_by_date]
        ### Open containers urls for user to validate in chrome
        title_formated = duplicates_ordered_by_date[0].title.replace(' ', '-').replace(':', '-')
        webbrowser.get(chrome_path).open(f'www.zxdsxzxs.pl/{title_formated}')
        for duplicate in duplicates_ordered_by_date:
            webbrowser.get(chrome_path).open(duplicate.container_url)
            time.sleep(pause_between_opening_urls)
        user_input = input(
            'Please enter which url which you like to delete separeted by dash (-)\n'
            'For e.g. there will be 3 urls displated in 3 tabs if you want to delete\n'
            'first and the last one enter "1-3" :\n'
        )
        user_input_idx = [int(x)-1 for x in user_input.split('-')]
        current_title_rows_to_delete = [
            duplicates_ordered_by_date[idx].cell_row for idx in user_input_idx
        ]
        rows_to_delete.extend(current_title_rows_to_delete)
        print(f'Rows to delete: {rows_to_delete}')
        print()

    ask_for_permision_to_delete = input(
        f'Those rows will be deleted:\n{rows_to_delete}\n'
        'Enter "y" to delete or "n" to pause the program and delete duplicates manually\n')
    if ask_for_permision_to_delete == 'y':
        return rows_to_delete
    print('"manual_urls_check" stopped by a user. Manual check needed to be done on xlsx file.\n\n')
    return []


def validate_duplicates(duplicates, pause_between_opening_urls, xlsx_file_names, input_wb):
    if len(duplicates) > 0:
        duplicates_to_fast_check = create_list_for_fast_check(duplicates=duplicates)
        ### try to check rows with fast method which delete older duplicates
        #   and leave the newest one
        if input(
                '"fast_urls_check" will check all of the newest containers url are valid\n'
                'and delete the older ones.\n\n'
                '"manual_urls_check" will ask user to choose which urls he wants to delete.\n\n'
                'Press "y" for "fast_urls_check" or "n" to continue to "manual_urls_check":\n'
            ) =='y':
            print()
            rows_to_delete = fast_urls_check(
                duplicates_ordered_by_date_found=duplicates_to_fast_check,
                pause_between_opening_urls=pause_between_opening_urls,
            )
            if len(rows_to_delete) >= 1:
                return sorted(rows_to_delete)
        print()
        if input('Press "y" for manual_urls_check or "n" to pause program execution '
                'and delete duplicates manually\n'
                '"manual_urls_check" will ask user to choose which urls he wants to delete:\n'
            ) == 'y':
            print()
            rows_to_delete = manual_urls_check(
                duplicates=duplicates,
                pause_between_opening_urls=pause_between_opening_urls,
            )
            if len(rows_to_delete) >= 1:
                return sorted(rows_to_delete)
        print('Both methods failed. Manual check needed. \n Program will now stop its execution.')
        exit()
    else:
        print(
            '~~~~~~~~~ There are no duplicates in this file ~~~~~~~~~\n'
            'This file will be saved as checked with prefix:\n'
            '"CheckedForDuplicates___"\n'
            'Program will now exit...\n'
        )
        input_wb.save(
            xlsx_file_names['prefix_after_duplicates_check'] + xlsx_file_names['input_file']
        )
        exit()


def delete_duplicated_rows(ws, rows_to_delete):
    rows_deleted = []
    for row in reversed(rows_to_delete):
        ws.delete_rows(row, amount=1)
        rows_deleted.append(row)
    return rows_deleted



### load input xlsx file
stream = open('config.yaml', 'r')
xlsx_file_names = yaml.safe_load(stream)
input_wb = load_workbook(xlsx_file_names['input_file'])
input_ws = input_wb.active

### create dictionary with duplicates structured into dictionary like
#   {duplicated_url_shorthened_url; [1st_ContainerAttributes, 2nd_ContainerAttributes], etc.}
duplicates = find_duplicated_urls(
    input_ws=input_ws,
    place_holder=xlsx_file_names['place_holder']
)


### Ask user for pause between opening new urls 
pause_between_opening_urls = float(input(
    'Pleae enter how much do you want to wait between opening URLs:\n'
    'It needs to be a Integer(eg.: 2) or Float(eg.: 1.3)\n'
))
print(f'Pause set to: {pause_between_opening_urls}')
print()

### return rows corresponding to duplicated containers
rows_to_delete = validate_duplicates(
    duplicates=duplicates, 
    pause_between_opening_urls=pause_between_opening_urls,
    xlsx_file_names=xlsx_file_names,
    input_wb=input_wb,
)
input_ws_initial_max_rows = input_ws.max_row
### delete selected rows
deleted_rows = delete_duplicated_rows(ws=input_ws, rows_to_delete=rows_to_delete)

### check if every row was deleted
print('~~~~~~~~~ Those duplicates were found ~~~~~~~~~\n')
for key, title_duplicates in duplicates.items():
    print(title_duplicates[0].title)
    for duplicate in title_duplicates:
        print(
            duplicate.container_url,
            duplicate.date_found,
            duplicate.cell_row,
            duplicate.website_url,
            sep='\n', end='\n',
        )
    print()
if len(rows_to_delete) == len(deleted_rows):
    print(f'Program deleted:\n{deleted_rows}.\n'
        '.:Execution proceeded correctly:.\n'
        f'Rows before check: {input_ws_initial_max_rows}.\n'
        f'Rows after check: {input_ws.max_row}.\n'
        f'Rows deleted: {len(rows_to_delete)}.\n'
        '~~Program will now exit...~~'
    )

### Save checked file to a new .xlsx file
input_wb.save(
    xlsx_file_names['prefix_after_duplicates_check'] + xlsx_file_names['input_file']
)