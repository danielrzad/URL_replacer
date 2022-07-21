from pprint import pprint
from openpyxl import load_workbook
from openpyxl.styles import NamedStyle
from ordered_set import OrderedSet
from urllib.parse import urlparse


import yaml
import os


def get_domain_name(url):
    if '//www.' in url:
        url = url[url.index('://www.'):]
        return url[:url.index('.', 7)+1]
    else:
        url = url[url.index('://'):]
        return url[:url.index('.')+1]


def get_container_id(url):
    return url[url.rfind('/')+1:].replace('.html', '')


def get_shortened_url(url):
    ### This function is cutting http/s in front of url to be sure there won't be http/https missmatches
    ### and .com fragment at the end of the netloc in case a website changes their .extension  
    urlparse("scheme://netloc/path;parameters?query#fragment")
    parsed_url = urlparse(url)
    netlocc = parsed_url.netloc[:parsed_url.netloc.rfind('.')].replace('www.', '')
    pathh = parsed_url.path
    #parameterss = parsed_url.parameters
    queryy = parsed_url.query
    fragmentt = parsed_url.fragment
    return (netlocc + pathh + queryy + fragmentt)


def check_for_reportfile(url):
        if 'report_file' in url:
            print('Replace "report_file" phrase done to:   ', url)
            return url.replace('report_file?id=', '')
        return url


def check_if_every_pair_is_present(pairs_dict, path_to_final_ws):
    wb = load_workbook(path_to_final_ws)
    ws = wb.active
    urls_in_dict = 0
    urls_in_final_xlsx = OrderedSet()
    urls_found = 0
    possible_duplicates = 0
    for urls_list in pairs_dict.values():
        for url in urls_list:
            url_not_found = 1
            for idx in range(4, ws.max_row + 1):
                #print(ws[f'L{idx}'].value)
                if url in ws[f'L{idx}'].value:
                    if url not in urls_in_final_xlsx:
                        urls_in_final_xlsx.add(url)
                        urls_found += 1
                        url_not_found = 0
                        continue
                    else:
                        print('Possible duplicate found', url)
                        url_not_found = 0
                        urls_found += 1
                        possible_duplicates += 1
            if url_not_found == 1:
                print('URL not found', url)
            urls_in_dict += 1
    print('URLs_in_dict', urls_in_dict)
    print('URLs_found', urls_found)
    if urls_in_dict == urls_found:
        print('Presence test passed without any issues.')
    elif urls_in_dict < urls_found:
        print('Presence test passed, but there are probably duplicates in final file.')
    elif urls_in_dict > urls_found:
        print('Presence test not passed, some links from pairs_file are missing.')

### Load xlsx file names
stream = open('config.yaml', 'r')
xlsx_file_names = yaml.safe_load(stream)
data_file = xlsx_file_names['prefix_after_duplicates_check'] + xlsx_file_names['input_file']
pairs_file = xlsx_file_names['url_replacement_file']


### Load an xlsx file with data to be processed
data_workbook = load_workbook(data_file)
data_ws = data_workbook.active
### Load an xlsx file in which containers and related URLs are stored
pairs_workbook = load_workbook(pairs_file)
pairs_ws = pairs_workbook.active['A']


### Make a list with unique container names
def make_unique_containers_urls_list(data_ws):
    unique_containers_urls = OrderedSet()
    for container_url in data_ws['L'][3:]:
        unique_containers_urls.add(get_domain_name(container_url.value))
    unique_containers_urls_checked = {
        url for url in unique_containers_urls if int(input(f'Is {url} a correct container url? type 1 for yes and 0 for no.  '))
    }
    return unique_containers_urls_checked

unique_containers_urls = make_unique_containers_urls_list(data_ws=data_ws)
# unique_containers_urls = {
#                           '://www.keeplinks.', '://mirrorace.', '://multiup.', '://ouo.',
#                           '://peeplink.', '://shrinke.', '://link1s.', '://www.keepfile.',
#                           '://baffleswerv.', '://bit.',
#}

# print(unique_containers_urls)

### Parse containers and related URLs into a dictionary
### Dictionary schema:
### {container_id: {container_url: [url1, url2, etc...]}} 
pairs_dict = {}
prev_key = ''
for url in pairs_ws:
    url = url.value
    domain = get_domain_name(url)
    if domain in unique_containers_urls:
        pairs_dict[get_shortened_url(url)] = []
        prev_key = get_shortened_url(url)
    else:
        pairs_dict[prev_key].append(check_for_reportfile(url))

# pprint(pairs_dict)

### check for data_ws maximum rows to make sure that any row won't be overleap
rows_to_process = data_ws.max_row - 3
deleted_placeholders = 0
deleted_empty_containers = 0

### if there are any placeholders fill them below
### eg.: '://en.wikipedia'
placeholder = 'en.wikipedia'
date_style = NamedStyle(name='datetime', number_format='M.DD.YY HH:MM')
rows_skipped = 0
pairs_found_during_iteration = 0
current_row = 4
while current_row <= data_ws.max_row:
    print(current_row)
    container_url = data_ws[f'L{current_row}'].value
    container_url_without_hhtp = get_shortened_url(container_url)
    if placeholder in container_url:
        data_ws.delete_rows(current_row)
        deleted_placeholders += 1
        continue
    elif container_url_without_hhtp in pairs_dict:
        pairs_found_during_iteration += 1
        number_of_host_urls = len(pairs_dict[container_url_without_hhtp])
        data_ws[f'L{current_row}'] = pairs_dict[container_url_without_hhtp][0]
        data_ws[f'I{current_row}'].style = date_style
        if number_of_host_urls > 1:
            values_to_copy = [cell.value for cell in data_ws[current_row:current_row]]
            data_ws.insert_rows(current_row+1, number_of_host_urls-1)
            for idx in range(1, number_of_host_urls):
                current_idx = current_row + idx
                for i, cell in enumerate(data_ws[current_idx]):
                    cell.value = values_to_copy[i]
                data_ws[f'L{current_idx}'] = pairs_dict[container_url_without_hhtp][idx]
                data_ws[f'I{current_idx}'].style = date_style
            current_row += number_of_host_urls
            continue
        current_row += 1
        continue
    else:
        if get_domain_name(container_url) in unique_containers_urls:
            data_ws.delete_rows(current_row)
            deleted_empty_containers += 1
            continue
    current_row += 1
    rows_skipped += 1
    print('Row skipped:    ', container_url)

### save processed file
final_folder = 'Final file' 
if not os.path.exists(final_folder):
    os.mkdir(final_folder)
data_workbook.save(f'{final_folder}/{data_file}')


### make a check if there are any pairs skipped
check_if_every_pair_is_present(pairs_dict=pairs_dict, path_to_final_ws=f'{final_folder}/{data_file}')


###check if every row was processed
pairs_found = len(pairs_dict)
if pairs_found + deleted_placeholders + deleted_empty_containers + rows_skipped == rows_to_process:
    print('File processed without any errors.')
else:
    print('File not processed correctly.')
print('Rows_to_process:    ', rows_to_process)
print('Deleted_placeholders:    ', deleted_placeholders)
print('Deleted_empty_containers:    ', deleted_empty_containers)
print('Row skipped:    ', rows_skipped)
print('Pairs_found:    ', pairs_found)
print('Rows_processed', deleted_placeholders + deleted_empty_containers + rows_skipped + pairs_found)

print('pairs_found_during_iteration', pairs_found_during_iteration)
print('pairs_found', pairs_found)
